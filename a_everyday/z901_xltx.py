import pandas as pd
import openpyxl
import os
import chardet

# ファイルパス
txt_file = r'C:\Users\sem3171\OneDrive - 株式会社三社電機製作所\デスクトップ\202502\pc外注転送\ZM122_P100_1220.TXT'
excel_file = r'C:\Users\sem3171\OneDrive - 株式会社三社電機製作所\デスクトップ\202502\pc外注転送\00_転送対象.xlsx'  # xltx → xlsx に変更

# エンコーディングの自動検出
with open(txt_file, 'rb') as f:
    result = chardet.detect(f.read())
    encoding = result['encoding']

# テキストファイルの読み込み
df = pd.read_csv(txt_file, sep='\t', encoding=encoding)  # タブ区切り
df['日付'] = pd.to_datetime(df['MDEZ_DAT00'], format='%Y%m%d')
df['年月'] = df['日付'].dt.strftime('%Y/%m')

# ユニークな年月リストを取得し、フィールド名として利用
unique_months = sorted(df['年月'].unique())

# 転送/所要の区分
df['転送'] = df['MDEZ_DELB0'].apply(lambda x: 1 if x == '転送' else 0)
df['所要'] = df['MDEZ_DELB0'].apply(lambda x: 1 if x in ['従所要', '入出予'] else 0)

# 集計（年月ごとに転送・所要の数を計算）
summary = df.groupby('年月').agg({'転送': 'sum', '所要': 'sum'}).reindex(unique_months, fill_value=0)
summary.loc['転送合計'] = summary['転送'].sum()
summary.loc['所要合計'] = summary['所要'].sum()

# Excelファイルの操作
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # O列（15列目）以降のデータをクリア
    for col in range(15, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col, value=None)

    # 新しいフィールド名をO列以降にセット
    start_col = 15  # O列
    col = start_col

    for month in unique_months:
        ws.cell(row=1, column=col, value=f'{month} 転送')
        ws.cell(row=2, column=col, value=summary.at[month, '転送'])
        col += 1
    
    ws.cell(row=1, column=col, value='転送合計')
    ws.cell(row=2, column=col, value=summary.at['転送合計', '転送'])
    col += 1

    for month in unique_months:
        ws.cell(row=1, column=col, value=f'{month} 所要')
        ws.cell(row=2, column=col, value=summary.at[month, '所要'])
        col += 1
    
    ws.cell(row=1, column=col, value='所要合計')
    ws.cell(row=2, column=col, value=summary.at['所要合計', '所要'])
    
    # 値のみを貼り付ける（数式解除）
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=col):
        for cell in row:
            cell.value = cell.value  # 数式を解除し、値のみ保持
    
    wb.save(excel_file)
    print('Excelファイル更新完了')
else:
    print('Excelファイルが存在しません')